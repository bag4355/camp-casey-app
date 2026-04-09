from __future__ import annotations

from collections import defaultdict
from datetime import time
from pathlib import Path

from openpyxl import load_workbook

from camp_casey_app.domain.models import BusDataset, BusStop, BusStopSchedule, SourceReference
from camp_casey_app.ingest.common import excel_source
from camp_casey_app.utils.text import dedupe_keep_order, normalize_text, slugify
from camp_casey_app.utils.time import minutes_since_anchor


SERVICE_PROFILE_MAP = {
    "Mon-Fri": ("weekday", "Weekday"),
    "Weekend&USandTraining Holiday": ("weekend_us_training", "Weekend / US holiday / Training holiday"),
}


def _derive_stop_aliases(name: str) -> list[str]:
    aliases = [name]
    normalized = normalize_text(name)
    if "(" in name and ")" in name:
        inside = name.split("(", 1)[1].rsplit(")", 1)[0]
        aliases.extend([part.strip() for part in inside.split("/")])
        aliases.extend([part.strip() for part in inside.split(",")])
    replacements = {
        "opposite the bowling ctr": "bowling center opposite",
        "web": "web",
        "cac": "cac",
        "burger king": "burger king",
        "theater": "theater",
        "warrior's club": "warriors club",
        "casey lodge": "casey lodge",
        "tennis court": "tennis court",
        "thunder dfac": "thunder dfac",
        "bus terminal": "bus terminal",
    }
    for key, alias in replacements.items():
        if key in normalized:
            aliases.append(alias)
    return dedupe_keep_order(aliases)


def parse_bus_file(path: Path) -> BusDataset:
    workbook = load_workbook(path, data_only=True)
    stop_records: dict[str, BusStop] = {}
    schedule_sources: dict[tuple[str, str], list[SourceReference]] = defaultdict(list)
    schedule_times: dict[tuple[str, str], list[time]] = defaultdict(list)
    profile_times: dict[str, list[time]] = defaultdict(list)
    profile_first_seen_time: dict[str, time] = {}

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        service_profile, service_label = SERVICE_PROFILE_MAP.get(sheet, (slugify(sheet), sheet))
        for row in range(3, worksheet.max_row + 1):
            stop_number = worksheet.cell(row, 2).value
            stop_name = worksheet.cell(row, 3).value
            if not stop_number or not stop_name:
                continue

            canonical_name = str(stop_name).strip()
            stop_id = slugify(canonical_name)
            source = excel_source(
                file_name=path.name,
                label=f"{path.name} • {sheet} • row {row}",
                sheet_name=sheet,
                row=row,
                column=3,
                excerpt=canonical_name,
            )
            if stop_id not in stop_records:
                stop_records[stop_id] = BusStop(
                    stop_id=stop_id,
                    name=canonical_name,
                    aliases=_derive_stop_aliases(canonical_name),
                    stop_numbers=[int(stop_number)],
                    source_refs=[source],
                )
            else:
                stop = stop_records[stop_id]
                if int(stop_number) not in stop.stop_numbers:
                    stop.stop_numbers.append(int(stop_number))
                stop.source_refs.append(source)

            for column in range(4, worksheet.max_column + 1):
                value = worksheet.cell(row, column).value
                if not isinstance(value, time):
                    continue
                schedule_times[(service_profile, stop_id)].append(value)
                profile_times[service_profile].append(value)
                profile_first_seen_time.setdefault(service_profile, value)
                schedule_sources[(service_profile, stop_id)].append(
                    excel_source(
                        file_name=path.name,
                        label=f"{path.name} • {sheet} • row {row} col {column}",
                        sheet_name=sheet,
                        row=row,
                        column=column,
                        excerpt=f"{canonical_name} {value.strftime('%H:%M')}",
                    )
                )

    service_profile_start_times: dict[str, time] = {}
    for profile, values in profile_times.items():
        service_profile_start_times[profile] = profile_first_seen_time.get(profile) or min(values, key=lambda item: item.hour * 60 + item.minute)

    schedules: list[BusStopSchedule] = []
    for (profile, stop_id), values in schedule_times.items():
        anchor = service_profile_start_times[profile]
        ordered_times = sorted(values, key=lambda item: minutes_since_anchor(item, anchor))
        schedules.append(
            BusStopSchedule(
                stop_id=stop_id,
                service_profile=profile,
                departures=ordered_times,
                source_refs=schedule_sources[(profile, stop_id)],
            )
        )

    return BusDataset(
        route_id="hovey-casey-loop",
        route_name="Hovey / Casey Shuttle Loop",
        source_file=path.name,
        service_profile_labels={key: label for _, (key, label) in SERVICE_PROFILE_MAP.items()},
        service_profile_start_times=service_profile_start_times,
        stops=sorted(stop_records.values(), key=lambda stop: stop.stop_numbers[0]),
        schedules=sorted(schedules, key=lambda item: (item.service_profile, item.stop_id)),
    )
