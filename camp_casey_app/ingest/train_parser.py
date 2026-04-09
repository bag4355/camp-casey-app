from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from camp_casey_app.domain.models import TrainDataset, TrainDeparture, TrainProvider, TrainServiceSheet
from camp_casey_app.ingest.common import excel_source, generated_source
from camp_casey_app.utils.text import dedupe_keep_order, slugify


SHEET_KEYWORDS = {
    "평일": ("weekday", "Weekday"),
    "토요일": ("saturday", "Saturday"),
    "일요일": ("sunday", "Sunday"),
}


DESTINATION_ALIASES = {
    "인천": ["incheon"],
    "청량리": ["cheongnyangni"],
    "소요산": ["soyosan"],
    "광운대": ["gwangwoondae"],
}


def _parse_sheet_metadata(sheet_name: str) -> tuple[str, str, str]:
    service_key, service_label = "weekday", "Weekday"
    for keyword, values in SHEET_KEYWORDS.items():
        if keyword in sheet_name:
            service_key, service_label = values
            break
    direction_label = sheet_name
    if "(" in sheet_name and ")" in sheet_name:
        inner = sheet_name.split("(", 1)[1].rsplit(")", 1)[0]
        parts = [part.strip() for part in inner.split(",")]
        if len(parts) >= 2:
            direction_label = parts[1]
    return service_key, service_label, direction_label


def parse_train_file(path: Path) -> TrainDataset:
    workbook = load_workbook(path, data_only=True)
    source_refs = [generated_source(path.name, f"{path.name} • workbook")]
    bosan_sheets: list[TrainServiceSheet] = []

    destinations_seen: list[str] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        service_key, service_label, direction_label = _parse_sheet_metadata(sheet_name)
        departures: list[TrainDeparture] = []
        sheet_sources = [
            excel_source(
                file_name=path.name,
                label=f"{path.name} • {sheet_name}",
                sheet_name=sheet_name,
                row=2,
                column=2,
                excerpt=direction_label,
            )
        ]
        for row in range(3, worksheet.max_row + 1):
            departure_time = worksheet.cell(row, 2).value
            destination = worksheet.cell(row, 3).value
            if not departure_time or not destination:
                continue
            destinations_seen.append(str(destination))
            departures.append(
                TrainDeparture(
                    departure_time=departure_time,
                    destination=str(destination).strip(),
                    source_refs=[
                        excel_source(
                            file_name=path.name,
                            label=f"{path.name} • {sheet_name} • row {row}",
                            sheet_name=sheet_name,
                            row=row,
                            column=2,
                            excerpt=f"{departure_time.strftime('%H:%M')} {destination}",
                        )
                    ],
                )
            )
        bosan_sheets.append(
            TrainServiceSheet(
                provider_id="bosan",
                service_key=service_key,
                service_label=service_label,
                direction_label=direction_label,
                departures=departures,
                source_refs=sheet_sources,
            )
        )

    bosan_aliases = ["bosan", "보산", "보산역", "bosan station", "line 1", "1호선"]
    for destination, aliases in DESTINATION_ALIASES.items():
        if destination in destinations_seen:
            bosan_aliases.extend(aliases)

    bosan_provider = TrainProvider(
        provider_id="bosan",
        station_name="Bosan Station / 보산역",
        aliases=dedupe_keep_order(bosan_aliases),
        available=True,
        notes=[
            "Only the Bosan timetable file is connected in this version.",
            "Sheet selection is calendar weekday / Saturday / Sunday. No separate public-holiday sheet was provided.",
        ],
        sheets=bosan_sheets,
        source_refs=source_refs,
    )

    jihaeng_placeholder = TrainProvider(
        provider_id="jihaeng",
        station_name="Jihaeng Station / 지행역",
        aliases=["jihaeng", "지행", "지행역", "jihaeng station"],
        available=False,
        not_available_reason="Jihaeng timetable data has not been uploaded yet.",
        notes=["Provider scaffold exists, but no timetable file is connected yet."],
        sheets=[],
        source_refs=[generated_source("placeholder", "Jihaeng provider placeholder")],
    )

    return TrainDataset(providers=[bosan_provider, jihaeng_placeholder])
